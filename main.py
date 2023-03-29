#Excel Spread Sheet opener/changer
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def save_xl_file(filename):
    xlopener = xl.load_workbook(filename)
    sheet = xlopener['Sheet1']

    for row in range(2,sheet.max_row+1):
        cell_obj = sheet.cell(row,3)
        correct_price = cell_obj.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = correct_price
        xlopener.save("tx.xlsx")


    data = Reference(sheet, min_row=2, max_row=sheet.max_row,min_col=4,max_col =4)

    chart = BarChart()
    chart.add_data(data)
    sheet.add_chart(chart,'c8')
    xlopener.save("tx.xlsx")